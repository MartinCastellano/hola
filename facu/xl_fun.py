#!python3

import openpyxl,csv

from openpyxl.utils import get_column_letter, column_index_from_string







def xl_2py():
    
    bocaslista2 = []
    wb = openpyxl.load_workbook('/home/martin/trainings/facu/bocas.xlsx',data_only=True)    
    
    sheet = wb['hoja1']
    

    for row in range(1,sheet.max_row+1):
        
        bocaslista1 = []     
                           
        for column in range(1,sheet.max_column+1):            
            
            valor = sheet[get_column_letter(column)+str(row)].value
            
            bocaslista1.append(valor)
        
        Boca_N = BocaRegistro(A,B,C,D)
        bocaslista2.append(bocaslista1)
        
    print(bocaslista2)

    csvfile = open('/home/martin/trainings/facu/nuevo.csv','w',newline='')
            
    wr = csv.writer(csvfile,delimiter =' ',quoting=csv.QUOTE_MINIMAL)
    # csv.writer(f, ,quotechar =',',quoting=csv.QUOTE_MINIMAL)
    wr.writerow(bocaslista2)
    csvfile.close()         

        
        
xl_2py()


