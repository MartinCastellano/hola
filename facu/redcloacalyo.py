#!/usr/bin/env python3
import openpyxl,csv

from openpyxl.utils import get_column_letter, column_index_from_string

# Boca_N = BocaRegistro(A,B,C,D)
class BocaRegistro:
    def __init__(self,denominacion, x, y,z):
        self.denominacion = denominacion
        self.x = x
        self.y = y
        self.z = z

def xl_2py():
    
    bocaslista2 = []
    wb = openpyxl.load_workbook('/home/martin/trainings/facu/bocas.xlsx',data_only=True)    
    sheet=wb.active
    
    # sheet = wb['hoja1']
    

    for row in range(1,sheet.max_row+1):
        
        boca = []     
                           
        for column in range(1,sheet.max_column+1):            
            
            valor = sheet[get_column_letter(column)+str(row)].value
            
            boca.append(valor)# a la lista le agrego el valor 

        
        Boca_N = BocaRegistro('BR_'+str(boca[0]),boca[1],boca[2],boca[3]) #creo un objeto con las caraceteristicas y lo agrego a la lista2
        
        bocaslista2.append(Boca_N)

        print(Boca_N.denominacion,Boca_N.x,Boca_N.y,Boca_N.z)# aca imprime y te muestra los valores de cada objeto 
        
        
    # print(bocaslista2)

           

xl_2py()