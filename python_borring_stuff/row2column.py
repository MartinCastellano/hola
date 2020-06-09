#!python3
import openpyxl


listaa = []

nwe=openpyxl.load_workbook('/home/martin/Downloads/Practicas/seleniumTest/borring stuff/sdfsf.xlsx')

sheet = nwe.get_active_sheet()

nwe2= openpyxl.Workbook()

sheet2 = nwe2.get_active_sheet()

for i in range(1,sheet.max_row+1):

    for j in range(1,sheet.max_column+1):


        if sheet.cell(row=i,column=j).value != None:

            print(sheet.cell(row=i,column=j).value)
            sheet2.cell(row=j,column=i).value = sheet.cell(row=i,column=j).value
            print(sheet2.cell(row=i,column=j).value)

nwe2.save('/home/martin/Downloads/Practicas/seleniumTest/borring stuff/sdfsf2.xlsx')
