#! python3

import openpyxl,sys,os

from openpyxl.utils import get_column_letter

from openpyxl.styles import Font

from openpyxl.styles import NamedStyle

N = 0

if sys.argv[1:]:
    print ('el elmento es N : '+"".join(sys.argv[1:]))
    N = ''.join(sys.argv[1:])
   # os.mkfile('/home/martin/Downloads/Practicas/seleniumTest/borring stuff/nuevito.xlsx')
else:
    N =input('ingrese el valor de N: ')

    
nw = openpyxl.Workbook()
nw.get_sheet_names()
sheet = nw.get_active_sheet()

for i in range(1,(int(N)+1)):
    
    sheet['A'+str(i+1)]=i
    sheet['A'+str(i+1)].font=Font(bold=True)
    
    sheet[get_column_letter(i+1)+'1']=i
    sheet[get_column_letter(i+1)+'1'].font=Font(bold=True)



for i in range(1,sheet.max_row):
    
    for j in range(1,sheet.max_column):
        
        sheet[get_column_letter(j+1)+str(i+1)].value = (j)*(i)



nw.save('/home/martin/Downloads/Practicas/seleniumTest/borring stuff/hola.xlsx')

openpyxl.load_workbook('/home/martin/Downloads/Practicas/seleniumTest/borring stuff/hola.xlsx')


    